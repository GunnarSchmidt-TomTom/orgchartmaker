#!/usr/bin/python3 
from dataclasses import dataclass, replace
import csv
import sys
import os

@dataclass(frozen=True)
class Employee:
    """Class for employee data"""
    id: str
    name: str
    manager_id: str
    title: str
    nreports: int = None

    def with_reports(self, nreports):
        return replace(self, nreports=nreports)

UNKNOWN_EMP = Employee("", "", "", "")

def read_csv(csv_file):
    employee_map = {"" : UNKNOWN_EMP} # map of every employee_id to the employee object
    manager_map = {} # map of every employee to its manager
    
    with open(csv_file, 'r') as f:
        reader = csv.reader(f, delimiter=';')
        next(reader) # skip header row
        for row in reader:
            if len(row) < 4:
                continue  # skip rows with insufficient columns
            (id, name, manager_id, title) = tuple([ x.strip() for x in row[:4] ])
            employee = Employee(id, name, manager_id, title)
            employee_map[id] = employee
            manager_map[employee] = manager_id

    return { employee: employee_map[manager_id] for (employee, manager_id) in manager_map.items() if manager_id in employee_map }

def read_excel(excel_file):
    """Read employee data from Excel file (.xlsx)"""
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl library required for Excel files. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    
    employee_map = {"" : UNKNOWN_EMP}
    manager_map = {}
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        if ws.max_row < 2:
            print(f"Error: Excel file '{excel_file}' has no data rows.", file=sys.stderr)
            sys.exit(1)
        
        # Skip header row, start from row 2
        for row_num in range(2, ws.max_row + 1):
            try:
                # Read columns: Unique Identifier, Name, Reports To, Line Detail 1 (title)
                id_cell = ws.cell(row_num, 1).value
                name_cell = ws.cell(row_num, 2).value  
                manager_id_cell = ws.cell(row_num, 3).value
                title_cell = ws.cell(row_num, 4).value
                
                # Convert None to empty string and strip whitespace
                id = str(id_cell).strip() if id_cell is not None else ""
                name = str(name_cell).strip() if name_cell is not None else ""
                manager_id = str(manager_id_cell).strip() if manager_id_cell is not None else ""
                title = str(title_cell).strip() if title_cell is not None else ""
                
                if not id:
                    print(f"Warning: Row {row_num} has empty ID, skipping.", file=sys.stderr)
                    continue
                    
                employee = Employee(id, name, manager_id, title)
                employee_map[id] = employee
                manager_map[employee] = manager_id
                
            except Exception as e:
                print(f"Warning: Error processing row {row_num}: {e}", file=sys.stderr)
                continue
                
    except Exception as e:
        print(f"Error reading Excel file '{excel_file}': {e}", file=sys.stderr)
        sys.exit(1)

    return { employee: employee_map[manager_id] for (employee, manager_id) in manager_map.items() if manager_id in employee_map }

def read_data_file(file_path):
    """Automatically detect file type and read employee data"""
    if not os.path.exists(file_path):
        print(f"Error: File '{file_path}' not found.", file=sys.stderr)
        sys.exit(1)
    
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext == '.csv':
        return read_csv(file_path)
    elif file_ext in ['.xlsx', '.xls']:
        return read_excel(file_path)
    else:
        print(f"Error: Unsupported file type '{file_ext}'. Supported formats: .csv, .xlsx, .xls", file=sys.stderr)
        sys.exit(1)

def extract_org(reports):
    def extract_recursively(boss):
        org = {}
        for report, manager in reports.items():
            if manager == boss:
                report_org = extract_recursively(report)
                nreports = sum([x.nreports for x in report_org]) + len(report_org)
                org[report.with_reports(nreports)] = report_org

        return org

    root_managers = set([ employee for (employee, manager) in reports.items() if manager == UNKNOWN_EMP])
    root_organizations =  { m : extract_recursively(m) for m in root_managers }

    return {m.with_reports(sum([x.nreports for x in org])) : org for (m, org) in root_organizations.items()}

def render_pydot(org):
    import pydot

    graph = pydot.Dot(graph_type='digraph', layout="dot", rankdir= "LR", splines = "line", overlap="scale", ranksep="2", strict= True)

    def render_recursively(o,  m  = None):
        subgraph = pydot.Subgraph(rank="same")
        graph.add_subgraph(subgraph)
        for employee, team in o.items():
            label = "<<B>{}</B><BR/>{}{}>".format(employee.name, employee.title, "<BR/>({})".format(employee.nreports) if employee.nreports else "")
            node = pydot.Node(employee.id, label=label, shape="box", style = "filled, rounded", fillcolor= "lightblue" if team else "lightyellow")
            subgraph.add_node(node)

            if m:
                graph.add_edge(pydot.Edge(m, node))

            render_recursively(team, node)

    render_recursively(org)
    graph.write_png("out.png")

def render_ascii(org):
    from treelib import Node, Tree

    tree = Tree()
    def render_recursively(o, m = None):
        for report, team in o.items():
            tree.create_node("{} ({}){}".format(report.name, report.title, " -- {}".format(report.nreports) if report.nreports else ""), report.id, parent = m.id if m else None)
            render_recursively(team, report)

    render_recursively(org)
    tree.show()

def org_sizes(org):
    def render_recursively(o, m = None):
        for report, team in o.items():
            if(report.nreports > 0):
                print("{};{};{};{}".format(report.name, report.title, report.nreports, len(team)))
                render_recursively(team, report)

    print("manager;title;total org size;number of directs")
    render_recursively(org)

def find_managers_org(org, root_manager):
    def find_root(o):
        for m, t in o.items():
            if m.id == root_manager or m.name == root_manager:
                yield {m : t}
            elif m.id > root_manager or m.name > root_manager:
                break
            else:
                for result in find_root(t):
                    yield result

    root_org = list(find_root(org))
    if len(root_org) != 1:
        raise Exception("Expected exactly one occurrence of {}, but found {}".format(root_manager, root_org))

    return root_org[0]


def only_managers(org):
    def om(o):
        return {employee : om(team) for employee, team in o.items() if employee.nreports > 0}
    
    return om(org)

def identity(org):
    return(org)

def main(args):
    reports = read_data_file(args.input)
    org = extract_org(reports)
    org = args.filter_org(org)

    if args.root:
        specific_org = find_managers_org(org, args.root[0])
        args.render(specific_org)
    else:
        args.render(org)

if __name__ == '__main__' :
    import argparse

    parser = argparse.ArgumentParser(description='Derive org tree from a CSV or Excel file.')
    parser.add_argument('--input', dest='input', required=True, 
                        help='Path to input file (.csv, .xlsx, or .xls)')
    parser.add_argument('--root', dest='root', nargs = 1, required=False, default=None,
                        help="Specifies an element in the tree as root, only that subtree will be rendered")

    parser.add_argument('--only-managers', dest='filter_org', action='store_const',
                        const=only_managers, default=identity,
                        help='will render the managers only')


    parser.add_argument('--ascii', dest='render', action='store_const',
                    const=render_ascii, default=render_pydot,
                    help='render an ascii tree')

    parser.add_argument('--orgsize', dest='render', action='store_const',
                    const=org_sizes, default=render_pydot,
                    help='outputs a CSV table showing the orgsize and number of direct reports')

    parser.add_argument('--pydot', dest='render', action='store_const',
                        const=render_pydot, default=render_pydot,
                        help='render into pydot image')

    main(parser.parse_args())
